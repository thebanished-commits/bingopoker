import pandas as pd
import openpyxl
import os
import io
import shutil
import requests

GOOGLE_SHEET_ID = "1LSb_nVlUkh6BpgdpAUneCpv5P4y65hTlyQFTLnD9jc4"
URL_POS_CSV = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/gviz/tq?tqx=out:csv&sheet=Posiciones"
URL_RAKE_CSV = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/gviz/tq?tqx=out:csv&sheet=rake"
URL_CAL_CSV = f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/gviz/tq?tqx=out:csv&sheet=Calendario"

LOCAL_EXCEL = r'C:\programas\poker\bingo poker club 2026.xlsx'
REL_EXCEL = os.path.join(os.path.dirname(__file__), 'bingo poker club 2026.xlsx')

if os.path.exists(LOCAL_EXCEL):
    EXCEL_PATH = LOCAL_EXCEL
else:
    EXCEL_PATH = REL_EXCEL

POINTS_RULES = {
    "Torneo Regular Shorthanded (SH)": {1: 30, 2: 10, 3: 0, 4: 0, 5: 0},
    "Torneo Regular Full ring (FR)": {1: 50, 2: 35, 3: 20, 4: 10, 5: 0},
    "Torneo Regular Multitable (MM)": {1: 70, 2: 50, 3: 30, 4: 20, 5: 0},
    "Main Event Shorthanded (SH)": {1: 70, 2: 30, 3: 0, 4: 0, 5: 0},
    "Main Event Full ring (FR)": {1: 100, 2: 75, 3: 40, 4: 20, 5: 10},
    "Main Event Multitable (MM)": {1: 130, 2: 90, 3: 50, 4: 35, 5: 20},
}

FECHAS_STANDARD = [f"{i:02d}" for i in range(1, 11)]

DEFAULT_CALENDARIO = [
    {"fecha": "Fecha 3", "mes": "Agosto", "desc": "Jueves 20 de agosto de 2026", "next": True},
    {"fecha": "Fecha 4", "mes": "Septiembre", "desc": "Jueves 3 de septiembre de 2026", "next": False},
    {"fecha": "Fecha 5", "mes": "Octubre", "desc": "Jueves 1 de octubre de 2026", "next": False},
    {"fecha": "Fecha 6", "mes": "Octubre", "desc": "Jueves 15 de octubre de 2026", "next": False},
    {"fecha": "Fecha 7", "mes": "Noviembre", "desc": "Jueves 5 de noviembre de 2026", "next": False},
    {"fecha": "Fecha 8", "mes": "Noviembre", "desc": "Jueves 19 de noviembre de 2026", "next": False},
    {"fecha": "Fecha 9", "mes": "Diciembre", "desc": "Jueves 3 de diciembre de 2026", "next": False},
    {"fecha": "Fecha 10", "mes": "Diciembre", "desc": "Jueves 17 de diciembre de 2026", "next": False},
]
DEFAULT_EVENTO_FINAL = "Sabado 09 de Enero 2027 desde las 14 horas"

def create_backup():
    if os.path.exists(EXCEL_PATH):
        backup_path = EXCEL_PATH.replace('.xlsx', '_backup.xlsx')
        try:
            shutil.copy(EXCEL_PATH, backup_path)
            return backup_path
        except Exception as e:
            print("Backup warning:", e)
    return None

def load_data():
    create_backup()
    
    # Load Calendario from Google Sheets if available
    schedule_data = DEFAULT_CALENDARIO
    evento_final = DEFAULT_EVENTO_FINAL
    
    try:
        res_cal = requests.get(URL_CAL_CSV, timeout=10)
        if res_cal.status_code == 200:
            df_c = pd.read_csv(io.StringIO(res_cal.text))
            parsed_cal = []
            for idx, r in df_c.iterrows():
                f_val = r.get("Fecha") if "Fecha" in r else r.iloc[0]
                m_val = r.get("Mes") if "Mes" in r else r.iloc[1]
                d_val = r.get("Descripcion") if "Descripcion" in r else (r.get("Descripción") if "Descripción" in r else r.iloc[2])
                
                if pd.isna(f_val) or str(f_val).strip() == "":
                    if not pd.isna(d_val) and str(d_val).strip():
                        evento_final = str(d_val).strip()
                else:
                    f_num_str = str(int(float(f_val))) if isinstance(f_val, (int, float)) and not pd.isna(f_val) else str(f_val).strip()
                    f_name = f"Fecha {f_num_str}" if not str(f_num_str).startswith("Fecha") else f_num_str
                    parsed_cal.append({
                        "fecha": f_name,
                        "mes": str(m_val).strip() if not pd.isna(m_val) else "",
                        "desc": str(d_val).strip() if not pd.isna(d_val) else "",
                        "next": (len(parsed_cal) == 0)
                    })
            if parsed_cal:
                schedule_data = parsed_cal
    except Exception as e:
        print("Calendario fetch warning:", e)

    # Attempt loading Posiciones & Rake from Google Sheets
    try:
        res_pos = requests.get(URL_POS_CSV, timeout=15)
        res_rake = requests.get(URL_RAKE_CSV, timeout=15)
        
        if res_pos.status_code == 200 and res_rake.status_code == 200:
            df_raw = pd.read_csv(io.StringIO(res_pos.text), header=None)
            
            subheaders = [str(df_raw.iloc[2, 3 + i]) if not pd.isna(df_raw.iloc[2, 3 + i]) else "" for i in range(10)]
            
            players_data = []
            for r in range(2, len(df_raw)):
                row = df_raw.iloc[r]
                p_name = row[2]
                if pd.isna(p_name) or str(p_name).strip() == "" or str(p_name).strip().lower() in ["jugador", "pos.", "fecha"]:
                    continue
                
                pts = {}
                for col_i, f_name in enumerate(FECHAS_STANDARD):
                    val = row[4 + col_i]
                    try:
                        pts[f_name] = int(float(val)) if not pd.isna(val) else 0
                    except Exception:
                        pts[f_name] = 0
                
                total_pts = sum(pts.values())
                players_data.append({
                    "Pos": len(players_data) + 1,
                    "Jugador": str(p_name).strip(),
                    **pts,
                    "Total": total_pts
                })
                
            df_pos = pd.DataFrame(players_data)
            if not df_pos.empty:
                df_pos = df_pos.sort_values(by="Total", ascending=False).reset_index(drop=True)
                df_pos["Pos"] = range(1, len(df_pos) + 1)
                
            cols_order = ["Pos", "Jugador"] + FECHAS_STANDARD + ["Total"]
            for col in cols_order:
                if col not in df_pos.columns:
                    df_pos[col] = 0
            df_pos = df_pos[cols_order]

            df_rake_raw = pd.read_csv(io.StringIO(res_rake.text), header=None)
            rake_cols = [f"F{i}" for i in range(1, 11)]
            rake_dict = {}
            for i in range(1, 11):
                val = df_rake_raw.iloc[1, i] if len(df_rake_raw) > 1 else 0
                try:
                    rake_dict[f"F{i}"] = float(val) if not pd.isna(val) else 0.0
                except Exception:
                    rake_dict[f"F{i}"] = 0.0
                    
            total_rake = sum(rake_dict.values())
            
            camp_total = total_rake * 0.50
            mf_total = total_rake * 0.40
            gastos_mf = total_rake * 0.10
            
            payouts = [
                {"Pos": 1, "Campeonato": int(camp_total * 0.50), "Mesa Final": int(mf_total * 0.50)},
                {"Pos": 2, "Campeonato": int(camp_total * 0.30), "Mesa Final": int(mf_total * 0.30)},
                {"Pos": 3, "Campeonato": int(camp_total * 0.20), "Mesa Final": int(mf_total * 0.20)},
            ]
            
            return {
                "df_posiciones": df_pos,
                "fechas_headers": FECHAS_STANDARD,
                "subheaders": subheaders,
                "rake_dict": rake_dict,
                "total_rake": total_rake,
                "camp_total": camp_total,
                "mf_total": mf_total,
                "gastos_mf": gastos_mf,
                "payouts": payouts,
                "schedule_data": schedule_data,
                "evento_final": evento_final
            }
    except Exception as e:
        print("Google Sheets fetch warning, falling back to local Excel:", e)

    # Local Excel Fallback
    xl = pd.ExcelFile(EXCEL_PATH)
    df_pos_raw = xl.parse('Posiciones', header=None)
    subheaders = [str(x) if not pd.isna(x) else "" for x in df_pos_raw.iloc[3, 4:14].values]
    
    players_data = []
    for idx in range(4, len(df_pos_raw)):
        row = df_pos_raw.iloc[idx]
        pos = row[1]
        player = row[2]
        if pd.isna(player) or str(player).strip() == "":
            continue
        
        pts = {}
        for col_i, f_name in enumerate(FECHAS_STANDARD):
            val = row[4 + col_i]
            pts[f_name] = int(float(val)) if not pd.isna(val) else 0
            
        total_pts = sum(pts.values())
        players_data.append({
            "Pos": int(pos) if not pd.isna(pos) else len(players_data) + 1,
            "Jugador": str(player).strip(),
            **pts,
            "Total": total_pts
        })
        
    df_pos = pd.DataFrame(players_data)
    if not df_pos.empty:
        df_pos = df_pos.sort_values(by="Total", ascending=False).reset_index(drop=True)
        df_pos["Pos"] = range(1, len(df_pos) + 1)
        
    for f in FECHAS_STANDARD:
        if f not in df_pos.columns:
            df_pos[f] = 0
            
    cols_order = ["Pos", "Jugador"] + FECHAS_STANDARD + ["Total"]
    df_pos = df_pos[cols_order]

    df_rake_raw = xl.parse('rake', header=None)
    rake_cols = [f"F{i}" for i in range(1, 11)]
    rake_vals = [float(x) if not pd.isna(x) else 0.0 for x in df_rake_raw.iloc[2, 1:11].values]
    
    rake_dict = dict(zip(rake_cols, rake_vals))
    total_rake = sum(rake_vals)
    
    camp_total = total_rake * 0.50
    mf_total = total_rake * 0.40
    gastos_mf = total_rake * 0.10
    
    payouts = [
        {"Pos": 1, "Campeonato": int(camp_total * 0.50), "Mesa Final": int(mf_total * 0.50)},
        {"Pos": 2, "Campeonato": int(camp_total * 0.30), "Mesa Final": int(mf_total * 0.30)},
        {"Pos": 3, "Campeonato": int(camp_total * 0.20), "Mesa Final": int(mf_total * 0.20)},
    ]
    
    return {
        "df_posiciones": df_pos,
        "fechas_headers": FECHAS_STANDARD,
        "subheaders": subheaders,
        "rake_dict": rake_dict,
        "total_rake": total_rake,
        "camp_total": camp_total,
        "mf_total": mf_total,
        "gastos_mf": gastos_mf,
        "payouts": payouts,
        "schedule_data": schedule_data,
        "evento_final": evento_final
    }

APPS_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbzVOT95SPKWcWP0kyFSDissECcWkXzl3O_YTsbdpmjFa8AYoDd3qigw6Dbma8cX4onnog/exec"

def save_data(df_posiciones, rake_dict, subheaders_list=None):
    """
    Saves data to Google Sheets via Apps Script webhook (primary, works on cloud)
    and to local Excel as secondary (only when file exists locally).
    """
    success_sheets = False
    error_msg = None

    # --- PRIMARY: Google Sheets via Apps Script webhook ---
    try:
        players_payload = []
        for p_data in df_posiciones.to_dict('records'):
            fechas_pts = {f: int(p_data.get(f, 0)) for f in FECHAS_STANDARD}
            players_payload.append({
                "jugador": p_data['Jugador'],
                "puntos": fechas_pts,
                "total": int(p_data.get('Total', sum(fechas_pts.values())))
            })

        rake_payload = {f"F{i}": float(rake_dict.get(f"F{i}", 0)) for i in range(1, 11)}

        payload = {
            "action": "save",
            "posiciones": players_payload,
            "rake": rake_payload,
            "subheaders": subheaders_list or []
        }

        resp = requests.post(APPS_SCRIPT_URL, json=payload, timeout=20)
        if resp.status_code == 200:
            success_sheets = True
        else:
            error_msg = f"Apps Script respondió con código {resp.status_code}: {resp.text[:200]}"
    except Exception as e:
        error_msg = f"Error al conectar con Google Sheets: {e}"

    # --- SECONDARY: Local Excel (only when running locally) ---
    if os.path.exists(EXCEL_PATH):
        try:
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws_pos = wb['Posiciones']

            if subheaders_list:
                for c_idx, sub in enumerate(subheaders_list):
                    ws_pos.cell(row=4, column=5 + c_idx, value=sub if sub else None)

            for r in range(5, 25):
                ws_pos.cell(row=r, column=2, value=None)
                ws_pos.cell(row=r, column=3, value=None)
                for c in range(5, 15):
                    ws_pos.cell(row=r, column=c, value=None)
                ws_pos.cell(row=r, column=15, value=None)

            for row_idx, p_data in enumerate(df_posiciones.to_dict('records')):
                r_num = 5 + row_idx
                ws_pos.cell(row=r_num, column=2, value=row_idx + 1)
                ws_pos.cell(row=r_num, column=3, value=p_data['Jugador'])
                tot = 0
                for c_idx, f_name in enumerate(FECHAS_STANDARD):
                    v = p_data.get(f_name, 0)
                    ws_pos.cell(row=r_num, column=5 + c_idx, value=int(v) if v > 0 else None)
                    tot += v
                ws_pos.cell(row=r_num, column=15, value=tot)

            ws_rake = wb['rake']
            total_r = 0
            for c_idx, f_name in enumerate([f'F{i}' for i in range(1, 11)]):
                val = rake_dict.get(f_name, 0)
                ws_rake.cell(row=3, column=2 + c_idx, value=float(val) if val > 0 else None)
                total_r += val
            ws_rake.cell(row=3, column=12, value=total_r if total_r > 0 else None)
            ws_pos.cell(row=3, column=17, value=total_r)

            camp_tot = total_r * 0.50
            mf_tot = total_r * 0.40
            payout_dist = [(0.50, 0.50), (0.30, 0.30), (0.20, 0.20)]
            for idx, (p_camp, p_mf) in enumerate(payout_dist):
                ws_pos.cell(row=5 + idx, column=18, value=int(camp_tot * p_camp))
                ws_pos.cell(row=5 + idx, column=19, value=int(mf_tot * p_mf))

            wb.save(EXCEL_PATH)
        except Exception as e:
            print("Local Excel save warning:", e)

    # Return result tuple: (success, error_message)
    return success_sheets, error_msg
